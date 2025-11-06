#Câu 1
def LuuFile(path, data):
    with open(path, 'a', encoding='utf-8') as f:
        f.write(data + "\n")

def DocFile(path):
    arrProduct = []
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            data = line.strip()
            arr = data.split(';')
            arrProduct.append(arr)
    return arrProduct
def LuuFile(path,data):
 file=open(path,'a',encoding='utf-8')
 file.writelines(data)
 file.writelines("\n")
 file.close()
def DocFile(path):
 arrProduct=[]
 file=open(path,'r',encoding='utf-8')
 for line in file:
    data=line.strip()
    arr=data.split(';')
    arrProduct.append(arr)
 file.close()
 return arrProduct
masp=input("nhập mã SP:")
tensp=input("nhập tên sp:")
dongia=float(input("nhập giá:"))
line=masp+";"+tensp+";"+str(dongia)
LuuFile("database.txt",line)
dssp=DocFile("database.txt")
#print(dssp)
def XuatSanPham(dssp):
 for row in dssp:
    for element in row:
        print(element,end='\t')
        print()
    print()
XuatSanPham(dssp)
def SortSp(dssp):
 for i in range(len(dssp)):
    for j in range(len(dssp)):
        a=dssp[i]
        b=dssp[j]
 if a[2]>b[2]:
    dssp[i]=b
 dssp[j]=a
SortSp(dssp)
print("Sản phẩm sau khi sắp xếp giá:")
XuatSanPham(dssp)

#print(dssp)
def XuatSanPham(dssp):
 for row in dssp:
    for element in row:
        print(element,end='\t')
        print()
    print()
XuatSanPham(dssp)
def SortSp(dssp):
 for i in range(len(dssp)):
    for j in range(len(dssp)):
        a=dssp[i]
        b=dssp[j]
 if a[2]>b[2]:
    dssp[i]=b
 dssp[j]=a
SortSp(dssp)
print("Sản phẩm sau khi sắp xếp giá:")
XuatSanPham(dssp)

#Câu 2
#: Tạo dữ liệu mẫu lưu vào csdl_so.txt
def LuuFile(path,data):
 file=open(path,'a',encoding='utf-8')
 file.writelines(data)
 file.writelines("\n")
 file.close()
def DocFile(path):
 arrSo=[]
 file=open(path,'r',encoding='utf-8')
 for line in file:

     data=line.strip()
 arr=data.split(',')
 arrSo.append(arr)
 file.close()
 return arrSo

from XuLyFile import *
LuuFile("csdl_so.txt","-5,4,7,9,3,20")
LuuFile("csdl_so.txt","5,-4,37,-19,24,-21")
LuuFile("csdl_so.txt","15,9,0,-38,-3,15")
LuuFile("csdl_so.txt","5,-4,77,-9,3,-7")
LuuFile("csdl_so.txt","55,44,27")
LuuFile("csdl_so.txt","-50,26")

from XuLyFile import *
arrSo=DocFile("csdl_so.txt")
print(arrSo)
def XuatSoAmTrenMoiDong(arrSo):
 for row in arrSo:
    for element in row:
        number=int(element)
 if number<0:
    print(number,end='\t')
 print()
print("Các số âm trên mỗi dòng:")
XuatSoAmTrenMoiDong(arrSo)
#Câu 3
from xml.dom.minidom import parse
import xml.dom.minidom
# Mở file xml bằng minidom parser
DOMTree = xml.dom.minidom.parse("employees.xml")
collection = DOMTree.documentElement
# Lấy tất cả tag là employee
employees = collection.getElementsByTagName("employee")
# Duyệt vòng lặp để lấy toàn bộ dữ liệu ra
for employee in employees:
 tag_id = employee.getElementsByTagName('id')[0]
 id=tag_id.childNodes[0].data
 tag_name = employee.getElementsByTagName('name')[0]
 name=tag_name.childNodes[0].data
 print(id,'\t',name)


 #Câu 4
 import json
jsonString = '{ "ma":"114", "age":25, "ten":"Liễu Đình Thương Mẫn"}'
dataObject=json.loads(jsonString)
print(dataObject)
print("Mã=",dataObject["ma"])
print("Tên=",dataObject["ten"])
print("Tuổi=",dataObject["age"])

#Câu 5
pythonObject = {
 "ten": "Liễu Đình Thương Mẫn",
 "tuoi": 25,
 "ma": "114"
}
import json
pythonObject = {
 "ten": "Liễu Đình Thương Mẫn",
 "tuoi": 25,
 "ma": "114"
}
jsonString = json.dumps(pythonObject)
# the result is a JSON string:
print(jsonString)

#Câu 6
import csv
with open('datacsv.csv', newline='') as f: reader = csv.reader(f, delimiter=';', quoting=csv.QUOTE_NONE)
for row in reader:
        print(row[1])


#Câu 7
import xlsxwriter
# Tạo một file excel cùng 1 sheet
workbook = xlsxwriter.Workbook('demo.xlsx')
worksheet = workbook.add_worksheet()
# thiết lập các cột cho file
worksheet.set_column('A:A', 5)
worksheet.set_column('B:B', 15)
worksheet.set_column('C:C', 20)
worksheet.set_column('D:D', 15)
worksheet.set_column('E:E', 15)
# định dạng tiêu đề cột in đậm
bold = workbook.add_format({'bold': True})
# thêm dòng tiêu đề và định dạng in đậm
worksheet.write('A1', 'STT',bold)
worksheet.write('B1', 'MÃ SẢN PHẨM',bold)
worksheet.write('C1', 'TÊN SẢN PHẨM',bold)
worksheet.write('D1', 'SỐ LƯỢNG',bold)
worksheet.write('E1', 'ĐƠN GIÁ',bold)
#thêm một dòng dữ liệu
worksheet.write('A2',1)
worksheet.write('B2','SP1')
worksheet.write('C2', 'Coca')
worksheet.write('D2', '15')
worksheet.write('E2', '15000')
#thêm một dòng dữ liệu
worksheet.write('A3',2)
worksheet.write('B3','SP2')
worksheet.write('C3', 'Pepsi')
worksheet.write('D3', '20')
worksheet.write('E3', '18000')
#Chèn Logo vào
worksheet.insert_image('B5', 'logo_UEL.png')
workbook.close()

#Câu 8
from openpyxl import load_workbook
wb = load_workbook('demo.xlsx')
print (wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.values:
 for value in row:
    print(value,"\t",end='')
 print("")

 #Câu 9
 import os

# ===============================
# LỚP ĐỐI TƯỢNG
# ===============================
class DanhMuc:
    def __init__(self, ma, ten):
        self.ma = ma
        self.ten = ten

class SanPham:
    def __init__(self, ma, ten, don_gia, ma_danh_muc):
        self.ma = ma
        self.ten = ten
        self.don_gia = float(don_gia)
        self.ma_danh_muc = ma_danh_muc


# ===============================
# DANH SÁCH LƯU TRONG BỘ NHỚ
# ===============================
danh_mucs = []
san_phams = []


# ===============================
# XỬ LÝ FILE
# ===============================
def luu_file():
    with open("danhmuc.txt", "w", encoding="utf-8") as f:
        for dm in danh_mucs:
            f.write(f"{dm.ma},{dm.ten}\n")
    with open("sanpham.txt", "w", encoding="utf-8") as f:
        for sp in san_phams:
            f.write(f"{sp.ma},{sp.ten},{sp.don_gia},{sp.ma_danh_muc}\n")
    print(" Đã lưu dữ liệu ra file!\n")

def doc_file():
    danh_mucs.clear()
    san_phams.clear()

    if os.path.exists("danhmuc.txt"):
        with open("danhmuc.txt", "r", encoding="utf-8") as f:
            for line in f:
                ma, ten = line.strip().split(",")
                danh_mucs.append(DanhMuc(ma, ten))

    if os.path.exists("sanpham.txt"):
        with open("sanpham.txt", "r", encoding="utf-8") as f:
            for line in f:
                ma, ten, don_gia, ma_danh_muc = line.strip().split(",")
                san_phams.append(SanPham(ma, ten, don_gia, ma_danh_muc))

    print(" Đã đọc dữ liệu từ file!\n")


# ===============================
# CHỨC NĂNG XỬ LÝ
# ===============================
def them_danh_muc():
    ma = input("Nhập mã danh mục: ")
    ten = input("Nhập tên danh mục: ")
    danh_mucs.append(DanhMuc(ma, ten))
    print(" Thêm danh mục thành công!\n")

def them_san_pham():
    ma = input("Nhập mã sản phẩm: ")
    ten = input("Nhập tên sản phẩm: ")
    don_gia = input("Nhập đơn giá: ")
    ma_dm = input("Nhập mã danh mục: ")
    san_phams.append(SanPham(ma, ten, don_gia, ma_dm))
    print(" Thêm sản phẩm thành công!\n")

def hien_thi():
    print("\n--- DANH SÁCH SẢN PHẨM ---")
    for sp in san_phams:
        dm = next((d for d in danh_mucs if d.ma == sp.ma_danh_muc), None)
        ten_dm = dm.ten if dm else "Không rõ"
        print(f"{sp.ma:10} | {sp.ten:20} | {sp.don_gia:10,.0f} | Danh mục: {ten_dm}")
    print()

def tim_kiem():
    tu_khoa = input("Nhập tên sản phẩm cần tìm: ").lower()
    ket_qua = [sp for sp in san_phams if tu_khoa in sp.ten.lower()]
    if ket_qua:
        for sp in ket_qua:
            print(f"{sp.ma:10} | {sp.ten:20} | {sp.don_gia:10,.0f}")
    else:
        print(" Không tìm thấy sản phẩm!\n")

def sua_san_pham():
    ma = input("Nhập mã sản phẩm cần sửa: ")
    for sp in san_phams:
        if sp.ma == ma:
            sp.ten = input(f"Tên mới ({sp.ten}): ") or sp.ten
            gia_moi = input(f"Đơn giá mới ({sp.don_gia}): ")
            if gia_moi:
                sp.don_gia = float(gia_moi)
            print(" Đã sửa sản phẩm!\n")
            return
    print(" Không tìm thấy sản phẩm!\n")

def xoa_san_pham():
    ma = input("Nhập mã sản phẩm cần xóa: ")
    for sp in san_phams:
        if sp.ma == ma:
            san_phams.remove(sp)
            print(" Đã xóa sản phẩm!\n")
            return
    print(" Không tìm thấy sản phẩm!\n")

def sap_xep():
    san_phams.sort(key=lambda x: x.don_gia)
    print(" Đã sắp xếp sản phẩm theo đơn giá tăng dần!\n")


# ===============================
# MENU CHÍNH
# ===============================
def menu():
    while True:
        print("""
===== QUẢN LÝ SẢN PHẨM =====
1. Thêm danh mục
2. Thêm sản phẩm
3. Hiển thị danh sách sản phẩm
4. Tìm kiếm sản phẩm
5. Sửa sản phẩm
6. Xóa sản phẩm
7. Sắp xếp sản phẩm
8. Lưu file
9. Đọc file
0. Thoát
============================
""")
        chon = input(" Nhập lựa chọn: ")
        if chon == "1": them_danh_muc()
        elif chon == "2": them_san_pham()
        elif chon == "3": hien_thi()
        elif chon == "4": tim_kiem()
        elif chon == "5": sua_san_pham()
        elif chon == "6": xoa_san_pham()
        elif chon == "7": sap_xep()
        elif chon == "8": luu_file()
        elif chon == "9": doc_file()
        elif chon == "0":
            print(" Tạm biệt!")
            break
        else:
            print(" Lựa chọn không hợp lệ!\n")


# ===============================
# CHẠY CHƯƠNG TRÌNH
# ===============================
if __name__ == "__main__":
    menu()

#Câu 10
import json
import os

# ===============================
# LỚP ĐỐI TƯỢNG
# ===============================
class Lop:
    def __init__(self, ma_lop, ten_lop):
        self.ma_lop = ma_lop
        self.ten_lop = ten_lop

class SinhVien:
    def __init__(self, ma_sv, ten_sv, nam_sinh, ma_lop):
        self.ma_sv = ma_sv
        self.ten_sv = ten_sv
        self.nam_sinh = int(nam_sinh)
        self.ma_lop = ma_lop


# ===============================
# DANH SÁCH LƯU TRONG BỘ NHỚ
# ===============================
ds_lop = []
ds_sv = []


# ===============================
# XỬ LÝ FILE JSON
# ===============================
def luu_file_json():
    data = {
        "lop": [vars(l) for l in ds_lop],
        "sinhvien": [vars(sv) for sv in ds_sv]
    }
    with open("sinhvien.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
    print(" Đã lưu dữ liệu ra file JSON!\n")

def doc_file_json():
    if not os.path.exists("sinhvien.json"):
        print(" Chưa có file dữ liệu để đọc!\n")
        return

    ds_lop.clear()
    ds_sv.clear()

    with open("sinhvien.json", "r", encoding="utf-8") as f:
        data = json.load(f)
        for l in data.get("lop", []):
            ds_lop.append(Lop(l["ma_lop"], l["ten_lop"]))
        for s in data.get("sinhvien", []):
            ds_sv.append(SinhVien(s["ma_sv"], s["ten_sv"], s["nam_sinh"], s["ma_lop"]))

    print(" Đã đọc dữ liệu từ file JSON!\n")


# ===============================
# CÁC CHỨC NĂNG
# ===============================
def them_lop():
    ma = input("Nhập mã lớp: ")
    ten = input("Nhập tên lớp: ")
    ds_lop.append(Lop(ma, ten))
    print(" Đã thêm lớp học!\n")

def them_sinh_vien():
    ma = input("Nhập mã sinh viên: ")
    ten = input("Nhập tên sinh viên: ")
    ns = input("Nhập năm sinh: ")
    ma_lop = input("Nhập mã lớp: ")
    ds_sv.append(SinhVien(ma, ten, ns, ma_lop))
    print(" Đã thêm sinh viên!\n")

def hien_thi():
    print("\n--- DANH SÁCH SINH VIÊN ---")
    for sv in ds_sv:
        lop = next((l for l in ds_lop if l.ma_lop == sv.ma_lop), None)
        tenlop = lop.ten_lop if lop else "Không rõ"
        print(f"{sv.ma_sv:10} | {sv.ten_sv:20} | {sv.nam_sinh:6} | Lớp: {tenlop}")
    print()

def tim_kiem():
    tu_khoa = input("Nhập tên sinh viên cần tìm: ").lower()
    ket_qua = [sv for sv in ds_sv if tu_khoa in sv.ten_sv.lower()]
    if ket_qua:
        for sv in ket_qua:
            print(f"{sv.ma_sv:10} | {sv.ten_sv:20} | {sv.nam_sinh}")
    else:
        print(" Không tìm thấy sinh viên!\n")

def sua_sinh_vien():
    ma = input("Nhập mã sinh viên cần sửa: ")
    for sv in ds_sv:
        if sv.ma_sv == ma:
            sv.ten_sv = input(f"Tên mới ({sv.ten_sv}): ") or sv.ten_sv
            ns = input(f"Năm sinh mới ({sv.nam_sinh}): ")
            if ns:
                sv.nam_sinh = int(ns)
            print(" Đã sửa thông tin sinh viên!\n")
            return
    print(" Không tìm thấy sinh viên!\n")

def xoa_sinh_vien():
    ma = input("Nhập mã sinh viên cần xóa: ")
    for sv in ds_sv:
        if sv.ma_sv == ma:
            ds_sv.remove(sv)
            print(" Đã xóa sinh viên!\n")
            return
    print(" Không tìm thấy mã sinh viên!\n")

def sap_xep():
    ds_sv.sort(key=lambda x: x.ten_sv.lower())
    print(" Đã sắp xếp danh sách sinh viên theo tên!\n")


# ===============================
# MENU CHÍNH
# ===============================
def menu():
    while True:
        print("""
===== QUẢN LÝ SINH VIÊN (JSON) =====
1. Thêm lớp học
2. Thêm sinh viên
3. Hiển thị danh sách sinh viên
4. Tìm kiếm sinh viên
5. Sửa thông tin sinh viên
6. Xóa sinh viên
7. Sắp xếp sinh viên
8. Lưu file JSON
9. Đọc file JSON
0. Thoát
====================================
""")
        chon = input(" Nhập lựa chọn: ")
        if chon == "1": them_lop()
        elif chon == "2": them_sinh_vien()
        elif chon == "3": hien_thi()
        elif chon == "4": tim_kiem()
        elif chon == "5": sua_sinh_vien()
        elif chon == "6": xoa_sinh_vien()
        elif chon == "7": sap_xep()
        elif chon == "8": luu_file_json()
        elif chon == "9": doc_file_json()
        elif chon == "0":
            print(" Tạm biệt!")
            break
        else:
            print(" Lựa chọn không hợp lệ!\n")


# ===============================
# CHẠY CHƯƠNG TRÌNH
# ===============================
if __name__ == "__main__":
    menu()

#Câu 11
import openpyxl
from openpyxl import Workbook, load_workbook
import os

# === Lớp Nhân viên ===
class NhanVien:
    def __init__(self, ma, ten, tuoi):
        self.ma = ma
        self.ten = ten
        self.tuoi = tuoi

# === Danh sách nhân viên ===
ds_nv = []

# === Thêm nhân viên mới ===
def them_nhan_vien():
    ma = input("Nhập mã nhân viên: ")
    ten = input("Nhập tên nhân viên: ")
    tuoi = int(input("Nhập tuổi nhân viên: "))
    ds_nv.append(NhanVien(ma, ten, tuoi))
    print(" Đã thêm nhân viên thành công!\n")

# === Hiển thị danh sách nhân viên ===
def hien_thi():
    if not ds_nv:
        print(" Chưa có nhân viên nào.")
        return
    print("{:<10} {:<15} {:<10}".format("MÃ", "TÊN", "TUỔI"))
    for nv in ds_nv:
        print("{:<10} {:<15} {:<10}".format(nv.ma, nv.ten, nv.tuoi))
    print()

# === Lưu vào Excel ===
def luu_excel(ten_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"
    ws.append(["STT", "Mã", "Tên", "Tuổi"])
    for i, nv in enumerate(ds_nv, start=1):
        ws.append([i, nv.ma, nv.ten, nv.tuoi])
    wb.save(ten_file)
    print(f" Đã lưu danh sách vào file {ten_file}\n")

# === Đọc từ Excel ===
def doc_excel(ten_file):
    global ds_nv
    if not os.path.exists(ten_file):
        print(" File không tồn tại!")
        return
    wb = load_workbook(ten_file)
    ws = wb.active
    ds_nv = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        stt, ma, ten, tuoi = row
        ds_nv.append(NhanVien(ma, ten, tuoi))
    print(f" Đã đọc {len(ds_nv)} nhân viên từ file {ten_file}\n")

# === Sắp xếp nhân viên theo tuổi tăng dần ===
def sap_xep_theo_tuoi():
    ds_nv.sort(key=lambda nv: nv.tuoi)
    print(" Đã sắp xếp nhân viên theo tuổi tăng dần!\n")

# === Menu chương trình ===
def menu():
    while True:
        print("===== QUẢN LÝ NHÂN VIÊN (Excel) =====")
        print("1. Thêm nhân viên")
        print("2. Hiển thị danh sách")
        print("3. Lưu vào file Excel")
        print("4. Đọc từ file Excel")
        print("5. Sắp xếp theo tuổi tăng dần")
        print("0. Thoát")
        chon = input(" Chọn chức năng: ")

        if chon == "1":
            them_nhan_vien()
        elif chon == "2":
            hien_thi()
        elif chon == "3":
            luu_excel("nhanvien.xlsx")
        elif chon == "4":
            doc_excel("nhanvien.xlsx")
        elif chon == "5":
            sap_xep_theo_tuoi()
        elif chon == "0":
            print(" Kết thúc chương trình!")
            break
        else:
            print(" Lựa chọn không hợp lệ!\n")

# === Chạy chương trình ===
if __name__ == "__main__":
    menu()


 #Câu 12
    import csv
import random

# === Hàm ghi file CSV ===
def ghi_file_csv(ten_file):
    with open(ten_file, mode="w", newline="") as file:
        writer = csv.writer(file, delimiter=';')
        for _ in range(10):  # 10 dòng
            dong = [random.randint(1, 100) for _ in range(10)]  # 10 số ngẫu nhiên
            writer.writerow(dong)
    print(f" Đã tạo file {ten_file} gồm 10 dòng dữ liệu ngẫu nhiên.\n")

# === Hàm đọc file CSV và tính tổng mỗi dòng ===
def doc_file_csv(ten_file):
    with open(ten_file, mode="r") as file:
        reader = csv.reader(file, delimiter=';')
        print(" Tổng giá trị mỗi dòng:")
        for i, dong in enumerate(reader, start=1):
            so_nguyen = [int(x) for x in dong if x.strip() != ""]
            tong = sum(so_nguyen)
            print(f"Dòng {i}: {dong} → Tổng = {tong}")
    print()

# === Chương trình chính ===
def main():
    ten_file = "dulieu_nhanvien.csv"
    while True:
        print("===== QUẢN LÝ DỮ LIỆU CSV =====")
        print("1. Tạo file CSV (10 dòng, mỗi dòng 10 số ngẫu nhiên)")
        print("2. Đọc file và tính tổng mỗi dòng")
        print("0. Thoát")
        chon = input(" Chọn chức năng: ")

        if chon == "1":
            ghi_file_csv(ten_file)
        elif chon == "2":
            doc_file_csv(ten_file)
        elif chon == "0":
            print(" Kết thúc chương trình!")
            break
        else:
            print(" Lựa chọn không hợp lệ!\n")

if __name__ == "__main__":
    main()


#Câu 13
import xml.etree.ElementTree as ET

# === Đọc dữ liệu từ file XML ===
def doc_nhom_thiet_bi(file):
    tree = ET.parse(file)
    root = tree.getroot()
    nhoms = []
    for n in root.findall("nhom"):
        ma = n.find("ma").text
        ten = n.find("ten").text
        nhoms.append({"ma": ma, "ten": ten})
    return nhoms

def doc_thiet_bi(file):
    tree = ET.parse(file)
    root = tree.getroot()
    thietbis = []
    for t in root.findall("thietbi"):
        ma = t.find("ma").text
        ten = t.find("ten").text
        manhom = t.attrib["manhom"]
        thietbis.append({"ma": ma, "ten": ten, "manhom": manhom})
    return thietbis


# === Chức năng hiển thị ===
def hien_thi_nhom(nhoms):
    print("\n===== DANH SÁCH NHÓM THIẾT BỊ =====")
    for n in nhoms:
        print(f"Mã nhóm: {n['ma']}, Tên nhóm: {n['ten']}")

def hien_thi_thiet_bi(thietbis):
    print("\n===== DANH SÁCH THIẾT BỊ =====")
    for t in thietbis:
        print(f"Mã TB: {t['ma']}, Tên TB: {t['ten']}, Mã nhóm: {t['manhom']}")

def hien_thi_theo_nhom(nhoms, thietbis):
    print("\n===== DANH SÁCH THIẾT BỊ THEO NHÓM =====")
    for n in nhoms:
        print(f"\n Nhóm {n['ten']} ({n['ma']}):")
        ds = [t for t in thietbis if t["manhom"] == n["ma"]]
        if ds:
            for t in ds:
                print(f"   - {t['ten']} (Mã TB: {t['ma']})")
        else:
            print("   (Không có thiết bị nào)")

def nhom_nhieu_thiet_bi_nhat(nhoms, thietbis):
    dem = {}
    for n in nhoms:
        dem[n["ma"]] = len([t for t in thietbis if t["manhom"] == n["ma"]])
    max_so = max(dem.values())
    print("\n===== NHÓM CÓ NHIỀU THIẾT BỊ NHẤT =====")
    for n in nhoms:
        if dem[n["ma"]] == max_so:
            print(f" {n['ten']} ({n['ma']}) có {max_so} thiết bị.")


# === CHƯƠNG TRÌNH CHÍNH ===
def main():
    file_nhom = "nhomthietbi.xml"
    file_tb = "thietbi.xml"

    nhoms = doc_nhom_thiet_bi(file_nhom)
    thietbis = doc_thiet_bi(file_tb)

    while True:
        print("\n========== QUẢN LÝ THIẾT BỊ (XML) ==========")
        print("1. Hiển thị danh sách Nhóm thiết bị")
        print("2. Hiển thị danh sách Thiết bị")
        print("3. Hiển thị thiết bị theo Nhóm")
        print("4. Xuất Nhóm có nhiều thiết bị nhất")
        print("0. Thoát")

        chon = input("👉 Chọn chức năng: ")
        if chon == "1":
            hien_thi_nhom(nhoms)
        elif chon == "2":
            hien_thi_thiet_bi(thietbis)
        elif chon == "3":
            hien_thi_theo_nhom(nhoms, thietbis)
        elif chon == "4":
            nhom_nhieu_thiet_bi_nhat(nhoms, thietbis)
        elif chon == "0":
            print(" Kết thúc chương trình.")
            break
        else:
            print(" Lựa chọn không hợp lệ, thử lại!\n")

if __name__ == "__main__":
    main()

